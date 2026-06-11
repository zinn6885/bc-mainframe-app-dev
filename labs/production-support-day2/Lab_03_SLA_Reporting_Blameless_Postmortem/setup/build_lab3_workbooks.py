#!/usr/bin/env python3
"""Build Lab 3 starter and solution Excel workbooks."""

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
bold = Font(bold=True)
header_fill = PatternFill("solid", fgColor="D9E1F2")

SAMPLE = [
    ("09:00", 1, 95, "Service up"),
    ("09:05", 1, 102, "Service up"),
    ("09:10", 0, None, "Service down"),
    ("09:15", 0, None, "Service down"),
    ("09:20", 1, 180, "Service up; slow response"),
    ("09:25", 1, 88, "Service up"),
]


def style_header_row(ws, row=1):
    for cell in ws[row]:
        if cell.value:
            cell.font = bold
            cell.fill = header_fill


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_workbook(filled: bool) -> Workbook:
    wb = Workbook()
    today = date.today().isoformat()

    ws = wb.active
    ws.title = "Raw Metrics"
    ws.append(["Timestamp", "ServiceHealth (1=up, 0=down)", "ResponseTimeMs", "Notes"])
    if filled:
        for ts, h, rt, note in SAMPLE:
            ws.append([ts, h, rt if rt is not None else "—", note])
    style_header_row(ws)
    set_col_widths(ws, [14, 28, 18, 30])

    ws2 = wb.create_sheet("SLA Calculations")
    ws2.append(["Timestamp", "ServiceHealth", "ResponseTimeMs", "Exceeds 150ms SLA?"])
    if filled:
        for ts, h, rt, _ in SAMPLE:
            exceeds = "—" if h == 0 else ("Yes" if rt and rt > 150 else "No")
            ws2.append([ts, h, rt if rt is not None else "—", exceeds])
    ws2.append([])
    ws2.append(["Metric", "Value"])
    ws2.append(["Downtime intervals (5 min each)", 2 if filled else ""])
    ws2.append(["Total downtime (minutes)", 10 if filled else ""])
    ws2.append(["Response SLA compliance %", "75%" if filled else ""])
    ws2.append(["Resolution SLA compliance %", "92%" if filled else ""])
    style_header_row(ws2)
    set_col_widths(ws2, [14, 16, 18, 22])

    ws3 = wb.create_sheet("MTTR MTBF")
    ws3.append(["Metric", "Formula", "Value"])
    for metric, formula, val in [
        ("Total time (minutes)", "Observation window", 60 if filled else ""),
        ("Downtime (minutes)", "From SLA Calculations", 10 if filled else ""),
        ("Uptime (minutes)", "Total time − downtime", 50 if filled else ""),
        ("Number of incidents", "Count downtime events", 1 if filled else ""),
        ("MTTR (minutes)", "Downtime ÷ incidents", 10 if filled else ""),
        ("MTBF (minutes)", "Uptime ÷ failures", 50 if filled else ""),
        ("Availability %", "(Total − downtime) ÷ total × 100", "83.33%" if filled else ""),
    ]:
        ws3.append([metric, formula, val])
    style_header_row(ws3)
    set_col_widths(ws3, [24, 32, 14])

    ws4 = wb.create_sheet("Availability")
    ws4.append(["Metric", "Value"])
    for metric, val in [
        ("Availability %", "83.33%" if filled else ""),
        ("SLA target (99.9% max downtime/month, min)", 43),
        ("This incident downtime (min)", 10 if filled else ""),
        ("Meets 99.9% if only incident?", "Yes" if filled else ""),
    ]:
        ws4.append([metric, val])
    style_header_row(ws4)
    set_col_widths(ws4, [42, 16])

    ws5 = wb.create_sheet("KPI Dashboard")
    kpi = [
        ("WEEKLY KPI DASHBOARD", ""),
        (f"Date: {today if filled else '_______________'}", ""),
        ("", ""),
        ("SERVICE HEALTH", ""),
        ("Service Availability", "83.33%" if filled else "_______ %"),
        ("Total Downtime (min)", 10 if filled else "_______"),
        ("Number of Incidents", 1 if filled else "_______"),
        ("", ""),
        ("SLA COMPLIANCE", ""),
        ("Response SLA", "75%" if filled else "_______ %"),
        ("Resolution SLA", "92%" if filled else "_______ %"),
        ("Overall SLA", "83.5%" if filled else "_______ %"),
        ("", ""),
        ("RELIABILITY METRICS", ""),
        ("MTTR (minutes)", 10 if filled else "_______"),
        ("MTBF (minutes)", 50 if filled else "_______"),
        ("", ""),
        ("TOP ALERTS", ""),
        ("Service Down", "1" if filled else "_______ times"),
        ("High Response Time", "1" if filled else "_______ times"),
        ("", ""),
        ("ACTION ITEMS", ""),
        ("1.", "Add port-conflict check to deployment runbook" if filled else ""),
        ("2.", "Enable CloudWatch alarm → on-call for P1" if filled else ""),
        ("3.", "Schedule postmortem review within 48h" if filled else ""),
    ]
    for a, b in kpi:
        ws5.append([a, b])
    ws5["A1"].font = bold
    set_col_widths(ws5, [28, 48])

    ws6 = wb.create_sheet("Postmortem")
    pm = [
        ["BLAMELESS POSTMORTEM", "Incident ID: INC-AWS-001"],
        ["DATE", today if filled else "_______________"],
        ["SERVICE", "payment-processor"],
        ["SEVERITY", "P1"],
        ["SLA MET?", "No — 10 min downtime in sample window" if filled else "Yes / No"],
        [""],
        ["WHAT HAPPENED?"],
        [
            "Rogue process held port 8080; payment-processor failed to start. CloudWatch ServiceHealth dropped to 0."
            if filled
            else "[Describe based on CloudWatch data and Lab 2 root cause]"
        ],
        [""],
        ["WHEN?"],
        ["Start time", "09:10" if filled else "_______________"],
        ["End time", "09:20" if filled else "_______________"],
        ["Duration (minutes)", 10 if filled else "_______________"],
        [""],
        ["5 WHYS"],
        ["1. Why did the service fail?", "Port 8080 already in use" if filled else ""],
        ["2. Why did that happen?", "rogue-process.py from Lab 2 setup" if filled else ""],
        ["3. Why did that happen?", "Not cleaned up after prior failure" if filled else ""],
        ["4. Why did that happen?", "No automated port check on service start" if filled else ""],
        ["5. Why did that happen?", "Runbook did not include ss/tulpn verification" if filled else ""],
        [
            "Root cause",
            "Port conflict from orphaned rogue process; missing pre-start port validation" if filled else "",
        ],
        [""],
        ["WHAT WENT WELL?"],
        ["•", "CloudWatch alarm detected downtime quickly" if filled else ""],
        [""],
        ["WHAT WENT WRONG?"],
        ["•", "No pre-start port validation on payment-processor" if filled else ""],
        [""],
        ["ACTION ITEMS"],
        ["Action", "Owner", "Due Date"],
        ["Add port check to systemd ExecStartPre", "Platform", "Next sprint" if filled else ""],
        ["CloudWatch ServiceDown alarm → on-call", "SRE", "This week" if filled else ""],
        [""],
        ["LESSONS LEARNED"],
        ["•", "Validate port availability before service restart" if filled else ""],
    ]
    for row in pm:
        ws6.append(row)
    ws6["A1"].font = bold
    set_col_widths(ws6, [22, 55, 16])

    ws7 = wb.create_sheet("Runbook")
    rb = [
        ["RUNBOOK / SOP", "Document ID: RB-AWS-001"],
        ["TITLE", "payment-processor Service Recovery"],
        ["TRIGGER", "CloudWatch Alarm: PaymentProcessor-ServiceDown; user reports transactions failing"],
        [""],
        ["STEP 1 — ACKNOWLEDGE", "Note time; check Lab3-SLA-Dashboard for impact duration"],
        ["STEP 2 — CHECK SERVICE STATUS", "systemctl status payment-processor"],
        ["STEP 3 — CHECK LOGS", "sudo journalctl -u payment-processor -n 50"],
        ["STEP 4 — CHECK PORT CONFLICT", "sudo ss -tulpn | grep 8080; pgrep -af rogue-process.py"],
        ["STEP 5 — KILL CONFLICTING PROCESS", "sudo kill -9 <PID>"],
        [
            "STEP 6 — RESTART SERVICE",
            "sudo systemctl reset-failed payment-processor; sudo systemctl restart payment-processor",
        ],
        [
            "STEP 7 — VERIFY",
            "systemctl is-active payment-processor; curl -s -o /dev/null -w '%{http_code}' http://localhost:8080",
        ],
        ["STEP 8 — CLOSE INCIDENT", "Document root cause; update resolution time; close ticket"],
        [""],
        ["ESCALATION"],
        ["L2 Support", "_______________"],
        ["L3 Support", "_______________"],
        ["On-call Manager", "_______________"],
    ]
    for row in rb:
        ws7.append(row)
    ws7["A1"].font = bold
    set_col_widths(ws7, [32, 62])

    return wb


def main():
    build_workbook(filled=False).save(ROOT / "template" / "lab3_starter.xlsx")
    build_workbook(filled=True).save(ROOT / "instructor" / "lab3_solution.xlsx")
    print("Built template/lab3_starter.xlsx and instructor/lab3_solution.xlsx")


if __name__ == "__main__":
    main()
